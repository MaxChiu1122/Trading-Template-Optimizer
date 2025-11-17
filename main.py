"""
Excel Trading Optimizer - Main Entry Point

This module provides the main workflow orchestration for the trading strategy
backtesting and optimization system. It coordinates between Excel configuration,
strategy execution, performance analysis, and visualization generation.

Key Features:
- Excel-driven configuration management
- Strategy backtesting with multi-period position support
- Rolling window parameter optimization
- Comprehensive performance analysis
- Automated visualization generation

Usage:
    python main.py                    # Run basic backtest
    main(optimize=True)              # Run parameter optimization
    main(optimize=False)             # Run single backtest

Author: Excel Trading Optimizer Project
License: MIT
"""

from excel_io import read_dashboard_inputs, write_results
from performance_metrics import calculate_performance_metrics
from generate_visuals import plot_visualization
from strategy import parse_strategy_logic, strategy_from_logic
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from optimizer import optimize_strategy
import pandas as pd


def insert_plot_into_excel(excel_path: str, image_path: str, sheet_name: str = "Visualization"):
    """Insert a plot image into the specified Excel sheet, replacing any existing images."""
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    # Remove existing images
    ws._images.clear()
    img = ExcelImage(image_path)
    img.anchor = 'A1'
    ws.add_image(img)
    wb.save(excel_path)


def main(optimize: bool = False, max_position: int = 3, excel_path: str = None, initial_cash: float = 6000):
    """Main entry point for running backtest or optimization workflow.
    
    Args:
        optimize: If True, run parameter optimization; if False, run single backtest
        max_position: Maximum position limit (default: 3). Caps positions at ±max_position (e.g., ±3 means max +3 long or -3 short)
        excel_path: Path to Excel template file (default: "excel/trading_template.xlsx")
        initial_cash: Initial capital for backtesting (default: 6000)
    """
    if excel_path is None:
        excel_path = "excel/trading_template.xlsx"
    # symbol = "ES=F"

    # Update market data and build indicators
    # update_excel_with_market_data(excel_path, symbol, download_data=False)

    # Read config and logic after market_data is updated
    try:
        config = read_dashboard_inputs(excel_path)
        config["excel_path"] = excel_path
    except FileNotFoundError:
        print(f"[ERROR] Excel file not found: {excel_path}")
        print("Please ensure the Excel template exists at the specified path.")
        return
    except Exception as e:
        print(f"[ERROR] Failed to read Excel configuration: {e}")
        return

    df = config["market_data"]
    print("[DEBUG] Columns in market_data after config:", df.columns.tolist())
    print(f"[INFO] Maximum Position Limit: ±{max_position}")
    if optimize:
        print("Running optimization mode...")
        results_df, test_trades_list, best_params_list, test_indicator_dfs = optimize_strategy(config, max_position=max_position)
        
        # Check if optimization returned empty results
        if results_df.empty or not test_trades_list:
            print("[ERROR] Optimization returned empty results. Check parameter ranges and strategy logic.")
            return
        
        all_trades = pd.concat(test_trades_list, ignore_index=True)

        # Find best parameter set by objective
        if not config.get("objective_weights"):
            print("[WARNING] No objective weights defined. Using first parameter set.")
            best_params = best_params_list[0] if best_params_list else {}
        else:
            best_metric = next(iter(config["objective_weights"].keys()))
            # Robust handling for all-NA or missing best_metric
            if best_metric not in results_df.columns:
                print(f"[ERROR] Metric '{best_metric}' not found in results_df columns: {results_df.columns.tolist()}")
                best_params = best_params_list[0] if best_params_list else {}
            elif results_df[best_metric].isna().all():
                print(f"[ERROR] All values for metric '{best_metric}' are NA. Cannot select best parameters.")
                best_params = best_params_list[0] if best_params_list else {}
            else:
                if config["objective_type"] == "MAX":
                    best_idx_label = results_df[best_metric].idxmax()
                else:
                    best_idx_label = results_df[best_metric].idxmin()
                if pd.isna(best_idx_label):
                    print(f"[ERROR] idxmax/idxmin returned NaN for metric '{best_metric}'. Using first parameter set as fallback.")
                    best_params = best_params_list[0] if best_params_list else {}
                else:
                    best_idx = results_df.index.get_loc(best_idx_label)
                    best_params = best_params_list[best_idx]
        
        # Update config with best parameters
        if best_params:
            config.update(best_params)
        from indicator_builder import build_indicators
        
        # --- Build test set data with correct indicators for metrics calculation ---
        # Use test_indicator_dfs which already have correct indicators for each test window
        param_cols = list(best_params_list[0].keys())
        test_rows = []
        for i, test_df in enumerate(test_indicator_dfs):
            if test_df is not None and not test_df.empty:
                # Add *_used columns for this window
                for k in param_cols:
                    test_df[f"{k}_used"] = best_params_list[i][k]
                test_rows.append(test_df)
        if test_rows:
            df_data = pd.concat(test_rows, ignore_index=True)
        else:
            df_data = pd.DataFrame()

        # Use only test set data for metrics calculation, ensuring we exclude training days
        # Filter to only dates where trades occurred (EntryDate or ExitDate)
        if not df_data.empty and not all_trades.empty:
            trade_dates = set(all_trades["EntryDate"].unique()) | set(all_trades["ExitDate"].unique())
            # Filter test set data to only include dates with trades
            test_set_market_data = df_data[df_data["Date"].isin(trade_dates)].copy()
        else:
            test_set_market_data = df_data.copy() if not df_data.empty else pd.DataFrame()
        
        # Fallback: if test_set_market_data doesn't have required columns, build from full dataset but filter to trade dates
        if test_set_market_data.empty or "Final_Price" not in test_set_market_data.columns:
            optimized_market_data = build_indicators(
                config["market_data"].copy(), 
                best_params, 
                builder_df=config.get("indicator_builder"),
                talib_df=config.get("talib_builder")
            )
            if not all_trades.empty:
                trade_dates = set(all_trades["EntryDate"].unique()) | set(all_trades["ExitDate"].unique())
                test_set_market_data = optimized_market_data[optimized_market_data["Date"].isin(trade_dates)].copy()
            else:
                test_set_market_data = optimized_market_data.copy()

        # Now calculate performance metrics with test set data only (excludes training days)
        # This ensures Accuracy only uses trading days data, not training days
        combined_metrics = calculate_performance_metrics(all_trades, test_set_market_data, initial_cash=initial_cash)
        write_results(excel_path, all_trades, combined_metrics)

        # Ensure all base columns are present by merging with original market data
        base_cols = ["Date", "Open", "High", "Low", "Close", "Volume", "Pt"]
        market_base = config["market_data"][base_cols].copy() if all(col in config["market_data"].columns for col in base_cols) else config["market_data"].copy()
        # Merge on Date, giving priority to test set values
        if not df_data.empty:
            df_data = pd.merge(df_data, market_base, on="Date", how="left", suffixes=("", "_mkt"))
            # For each base col, if missing in df_data, fill from market_base
            for col in base_cols:
                if col not in df_data.columns:
                    df_data[col] = df_data[f"{col}_mkt"]
            # Remove any *_mkt columns
            df_data = df_data[[c for c in df_data.columns if not c.endswith("_mkt")]]
        else:
            df_data = market_base.iloc[0:0].copy()

        # Reorder columns: Date, Open, High, Low, Close, Volume, Pt, indicators, *_used
        base_cols_present = [col for col in base_cols if col in df_data.columns]
        indicator_cols = [col for col in df_data.columns if col not in base_cols_present and not col.endswith("_used")]
        used_cols = [col for col in df_data.columns if col.endswith("_used")]
        ordered_cols = base_cols_present + indicator_cols + used_cols
        df_data = df_data[ordered_cols]

        # Visualization (use the test set rows)
        png_path = plot_visualization(df_data, all_trades, output_folder="images")
        insert_plot_into_excel(excel_path, png_path, sheet_name="Visualization")

        from excel_io import write_data_table
        # Write only the test set rows to the Data sheet
        write_data_table(excel_path, df_data, sheet_name="Data")
    else:
        # Run normal backtest
        from indicator_builder import build_indicators
        
        # Build indicators first before running strategy
        builder_df = config.get("indicator_builder")
        talib_df = config.get("talib_builder")
        params = config.get("param_map", {})
        df = build_indicators(df, params, builder_df=builder_df, talib_df=talib_df)
        
        logic_df = config["logic_table"]
        rule_dict = parse_strategy_logic(logic_df)
        # Pass parameters to strategy function for position sizing
        force_same_day_exit = config.get("force_same_day_exit", False)
        result_df = strategy_from_logic(df, rule_dict, params, max_position=max_position, force_same_day_exit=force_same_day_exit)
        
        if result_df.empty:
            print("[WARNING] No trades generated. Check your strategy logic and entry conditions.")
        
        metrics = calculate_performance_metrics(result_df, df, initial_cash)
        write_results(excel_path, result_df, metrics)
        
        try:
            png_path = plot_visualization(df, result_df, output_folder="images")
            insert_plot_into_excel(excel_path, png_path, sheet_name="Visualization")
        except Exception as e:
            print(f"[WARNING] Failed to generate visualization: {e}")

    # # Optional: open interactive HTML
    # if Path(html_path).exists():
    #     webbrowser.open(f"file://{Path(html_path).resolve()}")


if __name__ == "__main__":
    OPTIMIZE = True
    MAX_POSITION = 5
    
    main(optimize=OPTIMIZE, max_position=MAX_POSITION)